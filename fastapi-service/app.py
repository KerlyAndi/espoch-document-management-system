from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import os
import json
import asyncio
from typing import Optional, Dict, Any
import logging

# Document processing imports
try:
    import PyPDF2
    from docx import Document
    import openpyxl
except ImportError as e:
    print(f"Warning: Some document processing libraries not available: {e}")

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="ESPOCH Document Processing API",
    description="FastAPI service for processing academic documents",
    version="1.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify exact origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Pydantic models
class DocumentProcessRequest(BaseModel):
    documentId: int
    filePath: str
    options: Optional[Dict[str, Any]] = {}

class TextExtractionRequest(BaseModel):
    filePath: str
    documentType: Optional[str] = None

class SummarizeRequest(BaseModel):
    text: str
    maxLength: Optional[int] = 200

class ProcessingStatus(BaseModel):
    documentId: int
    status: str
    progress: int
    message: str
    result: Optional[Dict[str, Any]] = None

# In-memory storage for processing status (in production, use Redis or database)
processing_status: Dict[int, ProcessingStatus] = {}

@app.get("/")
async def root():
    return {
        "message": "ESPOCH Document Processing API",
        "version": "1.0.0",
        "status": "running",
        "endpoints": {
            "health": "/health",
            "process": "/process-document",
            "extract": "/extract-text",
            "summarize": "/summarize",
            "status": "/status/{doc_id}"
        }
    }

@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "service": "FastAPI Document Processor",
        "version": "1.0.0"
    }

@app.post("/process-document")
async def process_document(request: DocumentProcessRequest):
    """
    Process a document and extract metadata, text, and generate summary
    """
    try:
        document_id = request.documentId
        file_path = request.filePath
        
        logger.info(f"Processing document {document_id}: {file_path}")
        
        # Initialize processing status
        processing_status[document_id] = ProcessingStatus(
            documentId=document_id,
            status="processing",
            progress=0,
            message="Iniciando procesamiento..."
        )
        
        # Check if file exists
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Archivo no encontrado")
        
        # Update progress
        processing_status[document_id].progress = 25
        processing_status[document_id].message = "Extrayendo texto..."
        
        # Extract text based on file type
        extracted_text = await extract_text_from_file(file_path)
        
        # Update progress
        processing_status[document_id].progress = 50
        processing_status[document_id].message = "Analizando contenido..."
        
        # Generate metadata
        metadata = await generate_metadata(extracted_text, file_path)
        
        # Update progress
        processing_status[document_id].progress = 75
        processing_status[document_id].message = "Generando resumen..."
        
        # Generate summary
        summary = await generate_summary(extracted_text)
        
        # Final result
        result = {
            "documentId": document_id,
            "extractedText": extracted_text[:1000] + "..." if len(extracted_text) > 1000 else extracted_text,
            "textLength": len(extracted_text),
            "summary": summary,
            "metadata": metadata,
            "keywords": extract_keywords(extracted_text),
            "processedAt": "2024-01-01T00:00:00Z"
        }
        
        # Update final status
        processing_status[document_id] = ProcessingStatus(
            documentId=document_id,
            status="completed",
            progress=100,
            message="Procesamiento completado",
            result=result
        )
        
        logger.info(f"Document {document_id} processed successfully")
        return result
        
    except Exception as e:
        logger.error(f"Error processing document {document_id}: {str(e)}")
        
        # Update error status
        if document_id in processing_status:
            processing_status[document_id].status = "error"
            processing_status[document_id].message = f"Error: {str(e)}"
        
        raise HTTPException(status_code=500, detail=f"Error procesando documento: {str(e)}")

@app.get("/status/{document_id}")
async def get_processing_status(document_id: int):
    """
    Get the processing status of a document
    """
    if document_id not in processing_status:
        raise HTTPException(status_code=404, detail="Estado de procesamiento no encontrado")
    
    return processing_status[document_id]

@app.post("/extract-text")
async def extract_text_endpoint(request: TextExtractionRequest):
    """
    Extract text from a document file
    """
    try:
        if not os.path.exists(request.filePath):
            raise HTTPException(status_code=404, detail="Archivo no encontrado")
        
        extracted_text = await extract_text_from_file(request.filePath)
        
        return {
            "filePath": request.filePath,
            "extractedText": extracted_text,
            "textLength": len(extracted_text),
            "success": True
        }
        
    except Exception as e:
        logger.error(f"Error extracting text: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error extrayendo texto: {str(e)}")

@app.post("/summarize")
async def summarize_text(request: SummarizeRequest):
    """
    Generate a summary of the provided text
    """
    try:
        summary = await generate_summary(request.text, request.maxLength)
        
        return {
            "originalLength": len(request.text),
            "summary": summary,
            "summaryLength": len(summary),
            "success": True
        }
        
    except Exception as e:
        logger.error(f"Error generating summary: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generando resumen: {str(e)}")

# Helper functions
async def extract_text_from_file(file_path: str) -> str:
    """
    Extract text from various file formats
    """
    try:
        file_extension = os.path.splitext(file_path)[1].lower()
        
        if file_extension == '.pdf':
            return extract_text_from_pdf(file_path)
        elif file_extension in ['.doc', '.docx']:
            return extract_text_from_docx(file_path)
        elif file_extension in ['.xls', '.xlsx']:
            return extract_text_from_excel(file_path)
        elif file_extension == '.txt':
            return extract_text_from_txt(file_path)
        else:
            # Try to read as plain text
            return extract_text_from_txt(file_path)
            
    except Exception as e:
        logger.error(f"Error extracting text from {file_path}: {str(e)}")
        return f"Error extrayendo texto: {str(e)}"

def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF file"""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text.strip()
    except Exception as e:
        return f"Error leyendo PDF: {str(e)}"

def extract_text_from_docx(file_path: str) -> str:
    """Extract text from DOCX file"""
    try:
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    except Exception as e:
        return f"Error leyendo DOCX: {str(e)}"

def extract_text_from_excel(file_path: str) -> str:
    """Extract text from Excel file"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Hoja: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
            text += "\n"
        return text.strip()
    except Exception as e:
        return f"Error leyendo Excel: {str(e)}"

def extract_text_from_txt(file_path: str) -> str:
    """Extract text from TXT file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    except UnicodeDecodeError:
        try:
            with open(file_path, 'r', encoding='latin-1') as file:
                return file.read()
        except Exception as e:
            return f"Error leyendo archivo de texto: {str(e)}"
    except Exception as e:
        return f"Error leyendo archivo: {str(e)}"

async def generate_metadata(text: str, file_path: str) -> Dict[str, Any]:
    """
    Generate metadata for the document
    """
    file_stats = os.stat(file_path)
    
    return {
        "fileSize": file_stats.st_size,
        "wordCount": len(text.split()),
        "characterCount": len(text),
        "lineCount": len(text.split('\n')),
        "fileExtension": os.path.splitext(file_path)[1],
        "fileName": os.path.basename(file_path)
    }

async def generate_summary(text: str, max_length: int = 200) -> str:
    """
    Generate a simple extractive summary
    """
    if len(text) <= max_length:
        return text
    
    # Simple extractive summary - take first sentences up to max_length
    sentences = text.split('. ')
    summary = ""
    
    for sentence in sentences:
        if len(summary + sentence) <= max_length:
            summary += sentence + ". "
        else:
            break
    
    if not summary:
        summary = text[:max_length] + "..."
    
    return summary.strip()

def extract_keywords(text: str, max_keywords: int = 10) -> list:
    """
    Extract simple keywords from text
    """
    # Simple keyword extraction - most frequent words
    words = text.lower().split()
    
    # Filter out common Spanish stop words
    stop_words = {
        'el', 'la', 'de', 'que', 'y', 'a', 'en', 'un', 'es', 'se', 'no', 'te', 'lo', 'le',
        'da', 'su', 'por', 'son', 'con', 'para', 'al', 'del', 'los', 'las', 'una', 'como',
        'pero', 'sus', 'han', 'me', 'si', 'sin', 'sobre', 'este', 'ya', 'entre', 'cuando',
        'todo', 'esta', 'ser', 'son', 'dos', 'también', 'fue', 'había', 'era', 'muy'
    }
    
    # Count word frequency
    word_count = {}
    for word in words:
        word = word.strip('.,!?";()[]{}')
        if len(word) > 3 and word not in stop_words:
            word_count[word] = word_count.get(word, 0) + 1
    
    # Get most frequent words
    keywords = sorted(word_count.items(), key=lambda x: x[1], reverse=True)[:max_keywords]
    return [word for word, count in keywords]

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
